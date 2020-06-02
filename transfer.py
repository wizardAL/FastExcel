import openpyxl
import os
import lib.util as util
import lib.dict as dict
import lib.excel as excel

# 文件目录
DIR_PATH = "./"
SOURCE_TITLES = ['批号','序号','银行','交易时间','流水号','对方名称','金额']


# 主行数
def __main__():
  print('开始读取文件，请等待......')
  file_list = excel.get_filelist(DIR_PATH, 'xlsx')
  for filename in file_list:
    transferTo(filename)
  print('=========================================')


def transferTo(filename):
  wb = excel.open_file(filename)
  ws_origin = wb._sheets[0]
  ws_source = wb.create_sheet(title="文本格式")

  # 标题
  build_titles(ws_source) 

  title_row = -1
  source_row = 1
  for i, row in enumerate(excel.read_rows(ws_origin)):

    # 当前行数
    cur_row = i + 1

    # 判断是否是标题行
    if dict.is_head_first(ws_origin.cell(row=cur_row, column=1).value):
      title_row = cur_row
      continue

    # 忽略非数据行
    if cur_row <= title_row or title_row < 0: continue

    source_row = source_row + 1
    excel.append(ws_source, 1, source_row, util.now_format()) # 批号
    excel.append(ws_source, 2, source_row, source_row-1) # 序号
    excel.append(ws_source, 3, source_row, filename[0 : filename.find('.')]) # 银行

    
    for cell in row:
      origin_title = excel.get_title(ws_origin, cell.column, title_row)
      source_title = dict.get_trans_dic(origin_title)
      if source_title == None: continue
      
      # 拷贝cell
      column_index = SOURCE_TITLES.index(source_title) + 1
      value = excel.get_value(source_title, cell)
      excel.append(ws_source, column_index, source_row, value)
    
  wb.save(filename)
  print(filename + '保存成功...')


# 生成标题
def build_titles(ws):
  for i, title in enumerate(SOURCE_TITLES):
     ws.cell(column=i+1, row=1, value=title)


if __name__ == "__main__":
    __main__()