import openpyxl
import os
import lib.dict as dict
import lib.util as util
import lib.excel as excel
import lib.express as expre

INPUT_PATH = "./input.xlsx"
OUT_PATH = "./output.xlsx"
MAIN_ROW = 1


def write_sheet(wb, main_sheet, branch_sheet):
  '''
  # 读取数据写入到母版
  '''
  print("开始复制sheet: " + branch_sheet.title)
  main_titles = excel.get_titles(main_sheet)

  title_row = -1
  for i, row in enumerate(excel.read_rows(branch_sheet)):
    
    # 当前行数
    cur_row = i + 1

    # 判断是否是标题行
    if dict.is_head_first(branch_sheet.cell(row=cur_row, column=1).value):
      title_row = cur_row
      continue

    # 忽略非数据行
    if cur_row <= title_row or title_row < 0: continue

    # 遍历所有数据行
    flag = True
    for cell in row:
      if cell.value == None: continue
      branch_title = excel.get_title(branch_sheet, cell.column, title_row)
      _branch_title = dict.get_title_dic(branch_title) #从字典转化title
      
      # 处理函数表达式
      if expre.is_rule(_branch_title):
        cell_values = excel.get_cell_values(branch_sheet, cell, title_row)
        _branch_title = expre.exec_rule(_branch_title, cell_values)

      # 忽略不在母版列表的title
      if _branch_title not in main_titles: continue

      # 处理name
      global MAIN_ROW
      if flag :
        print('成功复制第' + str(MAIN_ROW) + '行数据......')
        flag = False
        MAIN_ROW = MAIN_ROW + 1
        clear_row(main_sheet, len(main_titles), MAIN_ROW) #清空源数据
        excel.append(main_sheet, 1, MAIN_ROW, branch_sheet.title)

      # 拷贝cell
      column_index = main_titles.index(_branch_title) + 1
      value = excel.get_value(_branch_title, cell)
      excel.append(main_sheet, column_index, MAIN_ROW, value)

# 清空row
def clear_row(sheet, col_num, row):
  for i in (1, col_num):
    sheet.cell(column=i, row=row, value=None)

# main methon
if __name__ == '__main__':
  print('开始读取文件，请等待......')
  try:
    wb = excel.open_file(INPUT_PATH)
    names = wb.sheetnames
    main_sheet = None
    for i, name in enumerate(names):
      if i == 0:
        continue
      elif i == 1:
        main_sheet = wb[name]
      else:
        write_sheet(wb, main_sheet, wb[name])
    wb.save(OUT_PATH)
    print('执行完成.')
  except Exception as e:
    print('系统发生异常：', e)
  os.system("pause")