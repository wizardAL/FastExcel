import openpyxl
import os
import lib.util as util


def open_file(path):
  '''
  打开文件
  '''
  return openpyxl.load_workbook(path)


def read_rows(sheet):
  '''
  sheet读取
  '''
  return tuple(sheet.rows)


def get_titles(sheet):
  '''
  获取所有单元格的标题
  '''
  list = []
  for i, row in enumerate(tuple(sheet.rows)):
    if i != 0: return list
    for cell in row:
      if cell.value == None: continue
      list.append(cell.value)
  return list


def get_filelist(dir, suffix):
  '''
  获取指定后缀的文件列表
  '''
  list = []
  for home, dirs, files in os.walk(dir):
    files = [f for f in files if not f[0] == '~']
    dirs[:] = [d for d in dirs if d not in ['input','build','dist']] # 忽略当前目录下的子目录
    for filename in files:
      if filename.endswith(suffix):
        list.append(filename)
  return list


def get_title(sheet, _column, _row):
  '''
  获取指定单元格的标题
  '''
  return sheet.cell(row=_row, column=_column).value


def append(sheet, col, row, val):
  '''
  写入cell
  '''
  sheet.cell(column=col, row=row, value= val)


def get_value(title, cell):
  '''
  获取cell值
  '''
  value = abs(cell.value) if util.is_negative(cell.value) else cell.value
  value = value[1: len(value)] if str(value).startswith('`') else value
  
  if value == None or value == '':
    return ""
  elif title == "时间":  
    return util.date_format(value) # 格式化时间
  elif (title == "收入" or title == "支出") and (isinstance(value, str)): 
    return float(value.replace(',', ''))  # 格式化金额
  else:
    return value


def get_cell_values(sheet, cell, title_row):
  '''
  获取当前cell所有的数据
  '''
  dict = {}
  col = cell.column
  for i in (1, col-1):
    title = get_title(sheet, i, title_row)
    value = get_title(sheet, i, cell.row)
    dict[title] = value
  return dict