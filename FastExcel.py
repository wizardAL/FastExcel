import openpyxl
import os
import shutil
import lib.dict as dict
import lib.util as util

INPUT_PATH = "./input.xlsx"
OUT_PATH = "./output.xlsx"
MAIN_ROW = 1


# 主函数
def main():
  print('开始读取文件，请等待......')
  try:
    wb = open_file(INPUT_PATH)
  except IOError:
    print('指定文件不存在，请检查以下文件是否存在：' + INPUT_PATH)
  else:
    names = wb.sheetnames
    main_sheet = None
    for i, name in enumerate(names):
      if i == 0:
        continue
      if i == 1:
        main_sheet = wb[name]
        continue
      write_sheet(wb, main_sheet, wb[name])
    wb.save(OUT_PATH)
    print('执行完成.')
    os.system('pause')


# 拷贝文件
def copy(input, output):
  validate(input)
  shutil.copyfile(input, output)
  return open_file(output)


# 打开文件
def open_file(path):
  return openpyxl.load_workbook(path)


# 验证数据格式
def validate(file):
  return True


# 读取数据写入到母版
def write_sheet(wb, main_sheet, branch_sheet):
  print("开始复制sheet: " + branch_sheet.title)
  main_titles = get_titles(main_sheet)

  title_row = 1
  for i, row in enumerate(read_rows(branch_sheet)):
    
    cur_row = i + 1

    if dict.is_head_first(branch_sheet.cell(row=cur_row, column=1).value):
      title_row = cur_row
      continue

    # ignore title
    if cur_row <= title_row: continue

    flag = True
    for cell in row:
      if cell.value == None: continue
      branch_title = get_title(branch_sheet, cell, title_row)
      _branch_title = dict.get_title_dic(branch_title) #从字典转化title
      if _branch_title not in main_titles: continue

      # 处理name
      global MAIN_ROW
      if flag :
        print('成功复制第' + str(MAIN_ROW) + '行数据......')
        flag = False
        MAIN_ROW = MAIN_ROW + 1
        append(main_sheet, 1, MAIN_ROW, branch_sheet.title)

      # 拷贝cell
      column_index = main_titles.index(_branch_title) + 1
      value = abs(cell.value) if util.is_negative(cell.value) else cell.value
      append(main_sheet, column_index, MAIN_ROW, value)



# sheet读取
def read_rows(sheet):
  return tuple(sheet.rows)


# 获取所有单元格的标题
def get_titles(sheet):
  list = []
  for i, row in enumerate(tuple(sheet.rows)):
    if i != 0: return list
    for cell in row:
      if cell.value == None: continue
      list.append(cell.value)
  return list


# 获取指定单元格的标题
def get_title(sheet, cell, _row):
  return sheet.cell(row=_row, column=cell.column).value

# 写入cell
def append(sheet, col, row, val):
  sheet.cell(column=col, row=row, value= val)


if __name__ == '__main__':
  main()
